#!/bin/env python3

# a python program to load all the csv files passed as arguments
# then merge them into a single csv file, also split the data into 
# groups demarked by the first 6 letters of the register number

import sys
import csv
from openpyxl import Workbook # make sure you have this installed. do pip install -r requirements.txt
from datetime import datetime

data =  {}
SECTION_LENGTH = 6
ROLL_LENGTH = 8

def getSection(roll):
    return roll[:SECTION_LENGTH]

for filename in sys.argv[1:]:
    with open(filename, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0] == 'id':
                continue
            if row[0] not in data:
                data[row[0]] = row[1]

entries = sorted(list(data.keys()))
sections = sorted({ getSection(data) for data in entries })

# calculate min time, max time and then find time taken
# timestamp in dict is of format: 2023-03-23 16:22:22.984742
timestamps = [datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S.%f') for timestamp in data.values()]
mintime = min(timestamps)
maxtime = max(timestamps)

# check if the mintime and maxtime are same date or not
if mintime.date() != maxtime.date():
    print('The data seems to span multiple days, please check the data.')
    exit(1)

filename = f'attendance-{datetime.strftime(mintime, "%Y-%m-%d")}.xlsx'

wb = Workbook()
wb.create_sheet('all')
wb.remove(wb['Sheet'])
[wb.create_sheet(section) for section in sections]

for entry in entries:
    section = getSection(entry)
    wb['all'].append([entry, data[entry]])
    wb[section].append([entry, data[entry]])

# set the column width of each sheet first column to fit the roll number
for sheet in wb:
    sheet.column_dimensions['A'].width = (ROLL_LENGTH + 2) * 1.2

# print analysis of the data
# total number of entries, number of sections, number of entries in each section, total time taken
print('Total number of entries: ', len(entries))
print('Number of sections: ', len(sections))
for section in sections:
    print('Number of entries in section ', section, ': ', len(list(wb[section].values)))
print('Started at: ', mintime)
print('Ended at: ', maxtime)
print('Time taken: ', maxtime - mintime)

wb.save(filename)
print('Saved to ', filename)